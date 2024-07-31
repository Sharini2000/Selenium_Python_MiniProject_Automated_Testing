import os
import sys
import pytest
import openpyxl
import logging
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
from pages.pages_job_search import Job_Search_Page as JSP

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))

@pytest.fixture(scope="module")
def setup():

   # driver = webdriver.Chrome(executable_path=r"C:\Users\shari\Downloads\chromedriver-win64\chromedriver-win64\chromedriver.exe")
    #driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    driver = webdriver.Chrome(ChromeDriverManager().install())
    workbook = openpyxl.load_workbook("TestData/Datasheet.xlsx")
    sheet = workbook['Testdata']
    total_rows = sheet.max_row
    total_columns = sheet.max_column
    for rows in range(1, total_rows + 1):
        if sheet.cell(rows,total_columns).value == "Y": # The last column in the datasheet has the run type. So this fetches data which has run type yes.
          job_title = sheet.cell(rows,1).value
          username = sheet.cell(rows,2).value
          password = sheet.cell(rows,3).value

    job_search_obj = JSP(driver, job_title, username, password)
    driver.get("https://www.simplyhired.ca/")
    driver.implicitly_wait(10)
    driver.maximize_window()
    yield driver, workbook, sheet, job_search_obj
    driver.quit()


def test_search_job(setup):
    driver, workbook, sheet, job_search_obj = setup

    try:
        jobs_numbers_count = job_search_obj.enter_job_title()
        assert jobs_numbers_count > 0, f"No jobs found for job title {job_search_obj.job_title}"
        logger.info(f"Test case for job title '{job_search_obj.job_title}' passed as expected with {jobs_numbers_count} jobs found.")
    except AssertionError as e:
        logger.error(f"Test case for job title '{job_search_obj.job_title}' failed: {e}")
        raise

def test_open_random_job(setup):
    driver, workbook, sheet, job_search_obj = setup  # Properly get the setup fixture values
    try:
        job_search_obj.select_job()  # Ensure this method exists in your class
        logger.info(f"Random job opened successfully.")
    except Exception as e:
        logger.error(f"Failed to open a random job : {e}")
        raise



















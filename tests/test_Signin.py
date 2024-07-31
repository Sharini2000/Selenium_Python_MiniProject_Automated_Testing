import os
import sys
import pytest
import openpyxl
import logging
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from pages.pages_signin import SignIn

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))

@pytest.fixture(scope="module")
def setup():
    #driver = webdriver.Chrome(executable_path=r"C:\Users\shari\Downloads\chromedriver-win64\chromedriver-win64\chromedriver.exe")
    #driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    workbook = openpyxl.load_workbook("TestData/Datasheet.xlsx")
    sheet = workbook['Testdata']
    total_rows = sheet.max_row
    total_columns = sheet.max_column
    for rows in range(1, total_rows + 1):
        if sheet.cell(rows,total_columns).value == "Y": # The last column in the datasheet has the run type. So this fetches data which has run type yes.
          job_title = sheet.cell(rows,1).value
          username = sheet.cell(rows,2).value
          password = sheet.cell(rows,3).value
    signin_obj = SignIn(driver, job_title, username, password)
    driver.get("https://www.simplyhired.ca/")
    driver.implicitly_wait(10)
    driver.maximize_window()
    yield driver, workbook, sheet, signin_obj
    driver.quit()

def test_signin(setup):
    driver, workbook, sheet,signin_obj = setup
    try:
        isSigninSuccess = signin_obj.SignIn_Process()
        assert isSigninSuccess, f"The SignIn Success PopUp does not appear.Therefore, SignIn failed."
        logger.info(f"Test case passed as expected with SignIn Success PopUp displayed")
    except AssertionError as e:
        logger.error(f"Test case failed with no SignIn Success PopUp displayed as expected: {e}")
        raise
